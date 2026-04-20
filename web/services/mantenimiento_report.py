"""Generación del Excel mensual de mantenimientos (reportados / realizados)."""

import io

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

from fastapi import HTTPException
from fastapi.responses import StreamingResponse

from core.constants import MESES_ES


def _require_openpyxl():
    if openpyxl is None:
        raise HTTPException(500, "Instala openpyxl: pip install openpyxl")


def build_mantenimiento_mes_excel(
    reportados: list,
    realizados: list,
    year: int,
    month: int,
    filename_prefix: str = "reporte_mantenimiento",
) -> StreamingResponse:
    _require_openpyxl()
    mes_lbl = f"{MESES_ES[month]} {year}"

    COLOR_HDR = "0A0E1A"
    COLOR_ACCENT = "00D4FF"
    COLOR_ROW = "1A2235"
    COLOR_WHITE = "FFFFFF"

    def hdr_font():
        return Font(bold=True, color=COLOR_ACCENT, size=10)

    def cell_font(color=COLOR_WHITE):
        return Font(color=color, size=10)

    def fill(hex_color: str):
        return PatternFill("solid", fgColor=hex_color)

    def border():
        s = Side(style="thin", color="1E2D45")
        return Border(left=s, right=s, top=s, bottom=s)

    def left_al():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    def center_al():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    wb = openpyxl.Workbook()

    def write_sheet(ws, title: str, headers: list, rows: list, col_widths: list):
        ws.title = title[:31]
        ws.sheet_view.showGridLines = False
        ws.append(headers)
        for ci, h in enumerate(headers, 1):
            c = ws.cell(1, ci)
            c.font = hdr_font()
            c.fill = fill(COLOR_HDR)
            c.border = border()
            c.alignment = center_al()
        for row in rows:
            ws.append(row)
            ri = ws.max_row
            for ci in range(1, len(headers) + 1):
                cell = ws.cell(ri, ci)
                cell.fill = fill(COLOR_ROW)
                cell.font = cell_font()
                cell.border = border()
                cell.alignment = left_al()
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 28

    ws1 = wb.active
    hdr1 = [
        "ID ticket",
        "Equipo",
        "Fecha apertura",
        "Fecha límite (SLA)",
        "Estado",
        "Título en GLPI",
    ]
    rows1 = [
        [
            r.get("ticket_id"),
            r.get("nombre"),
            r.get("fecha_apertura"),
            r.get("fecha_limite"),
            r.get("estado_txt"),
            r.get("titulo"),
        ]
        for r in reportados
    ]
    write_sheet(ws1, "Reportados mes", hdr1, rows1, [10, 26, 14, 14, 14, 44])

    ws2 = wb.create_sheet("Realizados mes")
    hdr2 = [
        "ID ticket",
        "Equipo",
        "Fecha cierre",
        "Fecha apertura",
        "Estado",
        "Título en GLPI",
        "Nota",
    ]
    rows2 = [
        [
            r.get("ticket_id"),
            r.get("nombre"),
            r.get("fecha_cierre"),
            r.get("fecha_apertura") or "—",
            r.get("estado_txt"),
            r.get("titulo"),
            r.get("nota") or "",
        ]
        for r in realizados
    ]
    write_sheet(ws2, "Realizados mes", hdr2, rows2, [10, 26, 14, 14, 14, 40, 24])

    ws3 = wb.create_sheet("Resumen")
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells("A1:B1")
    t1 = ws3.cell(1, 1, f"Informe de mantenimientos preventivos — {mes_lbl}")
    t1.font = Font(bold=True, color=COLOR_ACCENT, size=12)
    t1.fill = fill(COLOR_HDR)
    t1.alignment = left_al()
    ws3.row_dimensions[1].height = 26

    summary = [
        ("Tickets abiertos (reportados) en el mes", len(reportados)),
        ("Mantenimientos realizados (cerrados/resueltos) en el mes", len(realizados)),
    ]
    ri = 3
    for label, val in summary:
        ws3.cell(ri, 1, label).font = cell_font()
        ws3.cell(ri, 1).fill = fill(COLOR_ROW)
        ws3.cell(ri, 1).border = border()
        ws3.cell(ri, 1).alignment = left_al()
        ws3.cell(ri, 2, val).font = Font(bold=True, color=COLOR_WHITE, size=11)
        ws3.cell(ri, 2).fill = fill(COLOR_ROW)
        ws3.cell(ri, 2).border = border()
        ws3.cell(ri, 2).alignment = center_al()
        ri += 1
    ws3.column_dimensions["A"].width = 52
    ws3.column_dimensions["B"].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"{filename_prefix}_{year}-{month:02d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
