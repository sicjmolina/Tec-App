import io
from datetime import date

from fastapi import APIRouter
from fastapi import HTTPException
from fastapi.responses import StreamingResponse

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

from core.glpi_errors import glpi_http_error
from core.http_client import requests_lib as _requests
from core.state_meta import record_last_glpi_sync
from schemas import RenovacionConfirmarIn, RenovacionExcelCustomIn
from services.glpi import GLPIClient
from settings import get_merged_config

router = APIRouter(prefix="/api", tags=["renovation"])

_DATOS_PRUEBA_RENOVACION = [
    {"id": "1", "nombre": "PC-VENTAS-01", "estado": "Activo", "users_id": 101, "ram_mb": 4096, "disco_gb": 256, "tipo_disco": "HDD", "cpu": "Intel Core i3-6100", "fabricante": "DELL", "modelo": "OptiPlex 3050", "serial": "SN001", "usuario": "j.perez"},
    {"id": "2", "nombre": "PC-VENTAS-02", "estado": "Activo", "users_id": 102, "ram_mb": 8192, "disco_gb": 512, "tipo_disco": "HDD", "cpu": "Intel Core i5-7500", "fabricante": "HP", "modelo": "EliteDesk 800", "serial": "SN002", "usuario": "m.garcia"},
    {"id": "3", "nombre": "PC-CONTAB-01", "estado": "Activo", "users_id": 103, "ram_mb": 4096, "disco_gb": 128, "tipo_disco": "HDD", "cpu": "Intel Core i3-4170", "fabricante": "LENOVO", "modelo": "ThinkCentre M710", "serial": "SN003", "usuario": "a.lopez"},
    {"id": "4", "nombre": "PC-PRODUC-01", "estado": "Activo", "users_id": 104, "ram_mb": 16384, "disco_gb": 512, "tipo_disco": "SSD", "cpu": "Intel Core i7-9700", "fabricante": "HP", "modelo": "EliteDesk 880", "serial": "SN004", "usuario": "c.ramirez"},
    {"id": "5", "nombre": "PC-BODEGA-01", "estado": "Inactivo", "users_id": 0, "ram_mb": 16384, "disco_gb": 512, "tipo_disco": "SSD", "cpu": "Intel Core i5-10400", "fabricante": "DELL", "modelo": "OptiPlex 5080", "serial": "SN005", "usuario": ""},
    {"id": "6", "nombre": "PC-BODEGA-02", "estado": "Inactivo", "users_id": 0, "ram_mb": 32768, "disco_gb": 1024, "tipo_disco": "SSD", "cpu": "Intel Core i7-10700", "fabricante": "LENOVO", "modelo": "ThinkCentre M720", "serial": "SN006", "usuario": ""},
    {"id": "7", "nombre": "PC-DEVOL-01", "estado": "Inactivo", "users_id": 0, "ram_mb": 8192, "disco_gb": 256, "tipo_disco": "SSD", "cpu": "Intel Core i5-8400", "fabricante": "DELL", "modelo": "OptiPlex 7060", "serial": "SN007", "usuario": ""},
    {"id": "8", "nombre": "PC-RESERVA-01", "estado": "Inactivo", "users_id": 0, "ram_mb": 16384, "disco_gb": 512, "tipo_disco": "NVMe SSD", "cpu": "Intel Core i7-12700", "fabricante": "HP", "modelo": "EliteDesk 800 G8", "serial": "SN008", "usuario": ""},
]


def _num_from_any(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(",", ".")
    buff = []
    for ch in s:
        if ch.isdigit() or ch == ".":
            buff.append(ch)
        elif buff:
            break
    try:
        return float("".join(buff)) if buff else 0.0
    except Exception:
        return 0.0


def _mb_from_value(value, unit_hint: str = "") -> int:
    num = _num_from_any(value)
    hint = (unit_hint or "").upper()
    if num <= 0:
        return 0
    if "TB" in hint:
        return int(num * 1024 * 1024)
    if "GB" in hint:
        return int(num * 1024)
    if "KB" in hint:
        return int(num / 1024)
    if "MB" in hint:
        return int(num)
    return int(num)


def _gb_from_value(value, unit_hint: str = "") -> int:
    num = _num_from_any(value)
    hint = (unit_hint or "").upper()
    if num <= 0:
        return 0
    if "TB" in hint:
        return int(num * 1024)
    if "MB" in hint:
        return int(num / 1024)
    if "KB" in hint:
        return int(num / (1024 * 1024))
    if num > 10000:
        return int(num / 1024)
    return int(num)


def _is_ssd(tipo: str) -> bool:
    t = (tipo or "").upper()
    return "SSD" in t or "NVME" in t or "M.2" in t


def _score(ram_mb: int, disco_gb: int, tipo_disco: str) -> int:
    score = 0
    if ram_mb >= 32768:
        score += 45
    elif ram_mb >= 16384:
        score += 35
    elif ram_mb >= 8192:
        score += 20
    elif ram_mb >= 4096:
        score += 8
    if disco_gb >= 1000:
        score += 25
    elif disco_gb >= 512:
        score += 18
    elif disco_gb >= 256:
        score += 10
    elif disco_gb >= 128:
        score += 4
    if "NVME" in (tipo_disco or "").upper() or "M.2" in (tipo_disco or "").upper():
        score += 30
    elif _is_ssd(tipo_disco):
        score += 20
    return score


def _fmt_specs(ram_mb: int, disco_gb: int, tipo_disco: str, cpu: str) -> str:
    ram_str = f"{ram_mb // 1024} GB RAM" if ram_mb >= 1024 else f"{ram_mb} MB RAM"
    disco_str = f"{disco_gb} GB {tipo_disco or ''}".strip()
    cpu_short = (cpu or "—").split("@")[0].strip()
    return f"{ram_str} · {disco_str} · {cpu_short}"


def _parse_glpi_computers_full(client: GLPIClient, raw_list: list) -> list:
    ids = {str(c.get("id")) for c in raw_list if c.get("id")}
    if not ids:
        return []

    states = {str(s.get("id")): str(s.get("name") or "") for s in client._get_all("State", {}) if s.get("id") is not None}
    manufacturers = {str(m.get("id")): str(m.get("name") or "") for m in client._get_all("Manufacturer", {}) if m.get("id") is not None}
    models = {str(m.get("id")): str(m.get("name") or "") for m in client._get_all("ComputerModel", {}) if m.get("id") is not None}
    users = {str(u.get("id")): str(u.get("name") or "") for u in client._get_all("User", {"is_deleted": "0"}) if u.get("id") is not None}

    mem_items = client._get_all("Item_DeviceMemory", {})
    hdd_items = client._get_all("Item_DeviceHardDrive", {})
    cpu_items = client._get_all("Item_DeviceProcessor", {})
    device_hdd = {str(d.get("id")): str(d.get("designation") or "") for d in client._get_all("DeviceHardDrive", {}) if d.get("id") is not None}
    device_cpu = {str(d.get("id")): str(d.get("designation") or "") for d in client._get_all("DeviceProcessor", {}) if d.get("id") is not None}

    ram_by = {}
    disk_by = {}
    disk_type_by = {}
    cpu_by = {}

    for m in mem_items:
        if str(m.get("itemtype") or "").lower() != "computer":
            continue
        cid = str(m.get("items_id") or "")
        if cid not in ids:
            continue
        size_val = m.get("size") or m.get("capacity") or m.get("memory") or m.get("memorysize")
        unit = str(m.get("unit") or m.get("size_unit") or "MB")
        ram_by[cid] = ram_by.get(cid, 0) + _mb_from_value(size_val, unit)

    for d in hdd_items:
        if str(d.get("itemtype") or "").lower() != "computer":
            continue
        cid = str(d.get("items_id") or "")
        if cid not in ids:
            continue
        size_val = d.get("capacity") or d.get("size") or d.get("totalsize")
        unit = str(d.get("unit") or d.get("capacity_unit") or "GB")
        disk_by[cid] = disk_by.get(cid, 0) + _gb_from_value(size_val, unit)

        dev_id = str(d.get("deviceharddrives_id") or "")
        designation = (device_hdd.get(dev_id, "") or "").upper()
        txt = " ".join([designation, str(d.get("comment") or ""), str(d.get("designation") or "")]).upper()
        inferred = disk_type_by.get(cid, "")
        if "NVME" in txt or "M.2" in txt:
            disk_type_by[cid] = "NVMe SSD"
        elif "SSD" in txt and inferred != "NVMe SSD":
            disk_type_by[cid] = "SSD"
        elif not inferred:
            disk_type_by[cid] = "HDD"

    for p in cpu_items:
        if str(p.get("itemtype") or "").lower() != "computer":
            continue
        cid = str(p.get("items_id") or "")
        if cid not in ids or cid in cpu_by:
            continue
        dev_id = str(p.get("deviceprocessors_id") or "")
        cpu_name = device_cpu.get(dev_id) or str(p.get("designation") or "")
        cpu_by[cid] = cpu_name.strip()

    items = []
    for c in raw_list:
        cid = str(c.get("id") or "").strip()
        if cid not in ids:
            continue
        raw_uid = c.get("users_id")
        try:
            users_id_val = int(raw_uid) if raw_uid not in (None, "", False) else 0
        except (TypeError, ValueError):
            users_id_val = 0

        items.append({
            "id": cid,
            "nombre": str(c.get("name") or "").strip(),
            "estado": states.get(str(c.get("states_id")), str(c.get("states_id") or "")),
            "users_id": users_id_val,
            "ram_mb": int(ram_by.get(cid, 0)),
            "disco_gb": int(disk_by.get(cid, 0)),
            "tipo_disco": disk_type_by.get(cid, "N/D"),
            "cpu": cpu_by.get(cid, "N/D"),
            "fabricante": manufacturers.get(str(c.get("manufacturers_id")), "N/D"),
            "modelo": models.get(str(c.get("computermodels_id")), "N/D"),
            "serial": str(c.get("serial") or "").strip(),
            "usuario": users.get(str(c.get("users_id")), str(c.get("users_id") or "N/D")),
        })
    return items


def _analizar_renovacion(equipos: list) -> dict:
    UMBRAL_DEBIL = 30

    def _estado_cat(estado: str) -> str:
        s = (estado or "").strip().lower()
        if s == "1":
            return "activo"
        if s == "2":
            return "inactivo"
        if "inactivo" in s or "stock" in s or "reserva" in s or "almacen" in s or "almacén" in s:
            return "inactivo"
        if "activo" in s or "en uso" in s or "produccion" in s or "producción" in s:
            return "activo"
        if "baja" in s or "retir" in s or "obsole" in s:
            return "baja"
        return "otro"

    activos = [e for e in equipos if _estado_cat(e["estado"]) == "activo"]
    inactivos = [e for e in equipos if _estado_cat(e["estado"]) == "inactivo"]

    for e in equipos:
        e["score"] = _score(e["ram_mb"], e["disco_gb"], e["tipo_disco"])
        e["specs_fmt"] = _fmt_specs(e["ram_mb"], e["disco_gb"], e["tipo_disco"], e["cpu"])

    debiles = sorted([e for e in activos if e["score"] <= UMBRAL_DEBIL], key=lambda x: x["score"])
    candidatos = sorted([e for e in inactivos if e["score"] > UMBRAL_DEBIL], key=lambda x: x["score"], reverse=True)

    pares = []
    usados_inactivos = set()
    for d in debiles:
        mejor = next((c for c in candidatos if c["id"] not in usados_inactivos and c["score"] > d["score"]), None)
        pares.append({
            "activo": d,
            "reemplazo": mejor,
            "mejora_ram": (mejor["ram_mb"] - d["ram_mb"]) // 1024 if mejor else 0,
            "mejora_disco": (mejor["disco_gb"] - d["disco_gb"]) if mejor else 0,
            "mejora_ssd": (_is_ssd(mejor["tipo_disco"]) and not _is_ssd(d["tipo_disco"])) if mejor else False,
            "ganancia_score": (mejor["score"] - d["score"]) if mejor else 0,
        })
        if mejor:
            usados_inactivos.add(mejor["id"])

    return {
        "total_activos": len(activos),
        "total_inactivos": len(inactivos),
        "debiles": len(debiles),
        "candidatos": len(candidatos),
        "debiles_items": debiles,
        "inactivos_items": inactivos,
        "candidatos_items": candidatos,
        "pares": pares,
        "todos": equipos,
    }


@router.get("/renovacion/diagnostico")
def get_renovacion_diagnostico():
    cfg = get_merged_config()
    if not cfg.get("glpi_url"):
        raise HTTPException(400, "Configura la URL de GLPI primero.")
    if _requests is None:
        raise HTTPException(500, "Instala requests: pip install requests")

    g = GLPIClient(cfg)
    extra = {}
    raw = []
    try:
        g.login()
        raw = g._get_all("Computer", {"is_deleted": "0", "is_template": "0", "range": "0-2"})
        if raw:
            c0 = raw[0]
            extra = {
                "equipo_ejemplo_id": c0.get("id"),
                "equipo_ejemplo_nombre": c0.get("name"),
                "states_id": c0.get("states_id"),
                "users_id": c0.get("users_id"),
                "state_name_resuelto": g.get_state_name_by_id(c0.get("states_id")) if c0.get("states_id") else "",
                "user_name_resuelto": g.get_user_name_by_id(c0.get("users_id")) if c0.get("users_id") else "",
                "item_device_memory_sample": g.get_linked_items(c0, "Item_DeviceMemory")[:3],
                "item_device_harddrive_sample": g.get_linked_items(c0, "Item_DeviceHardDrive")[:3],
                "item_device_processor_sample": g.get_linked_items(c0, "Item_DeviceProcessor")[:3],
            }
    except Exception as ex:
        glpi_http_error(ex, "diagnóstico de renovación")
    finally:
        g.logout()

    record_last_glpi_sync()
    return {
        "muestra": raw[:3],
        "total_campos": [list(c.keys()) for c in raw[:3]],
        "diagnostico_specs": extra,
    }


@router.get("/renovacion")
def get_renovacion(modo_prueba: bool = True):
    if modo_prueba:
        equipos = [{**e} for e in _DATOS_PRUEBA_RENOVACION]
        for e in equipos:
            e["score"] = _score(e["ram_mb"], e["disco_gb"], e["tipo_disco"])
            e["specs_fmt"] = _fmt_specs(e["ram_mb"], e["disco_gb"], e["tipo_disco"], e["cpu"])
        return _analizar_renovacion(equipos)

    cfg = get_merged_config()
    if not cfg.get("glpi_url"):
        raise HTTPException(400, "Configura la URL de GLPI primero.")
    if _requests is None:
        raise HTTPException(500, "Instala requests: pip install requests")

    g = GLPIClient(cfg)
    try:
        g.login()
        raw = g.get_computers_full()
        equipos = _parse_glpi_computers_full(g, raw)
    except Exception as ex:
        glpi_http_error(ex, "cargar equipos para análisis de renovación")
    finally:
        g.logout()

    record_last_glpi_sync()
    if not equipos:
        raise HTTPException(502, "GLPI no devolvió equipos con especificaciones.")
    return _analizar_renovacion(equipos)


def _build_renovacion_excel_response(data: dict, filename_prefix: str = "renovacion_equipos"):
    if openpyxl is None:
        raise HTTPException(500, "Instala openpyxl: pip install openpyxl")

    pares = data.get("pares", [])
    wb = openpyxl.Workbook()

    COLOR_HDR = "0A0E1A"
    COLOR_ACCENT = "00D4FF"
    COLOR_DEBIL = "2D1F07"
    COLOR_WHITE = "FFFFFF"

    def hdr_font(color="FFFFFF"):
        return Font(bold=True, color=color, size=10)

    def cell_font(bold=False, color="E2E8F0"):
        return Font(bold=bold, color=color, size=10)

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def border():
        s = Side(style="thin", color="1E2D45")
        return Border(left=s, right=s, top=s, bottom=s)

    def center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def left():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws1 = wb.active
    ws1.title = "Reemplazos recomendados"
    ws1.sheet_view.showGridLines = False

    HDR1 = [
        "Equipo activo", "Usuario", "Serial", "Fabricante / Modelo",
        "Score actual", "Specs actuales",
        "→ Reemplazo sugerido", "Serial reemplazo",
        "Score reemplazo", "Specs reemplazo",
        "Mejora RAM (GB)", "Mejora Disco (GB)", "Gana SSD?", "Mejora score",
    ]
    ws1.append(HDR1)
    for ci, h in enumerate(HDR1, 1):
        cell = ws1.cell(1, ci)
        cell.font = hdr_font(COLOR_ACCENT)
        cell.fill = fill(COLOR_HDR)
        cell.border = border()
        cell.alignment = center()

    for p in pares:
        d = p["activo"]
        r = p["reemplazo"]
        row = [
            d["nombre"],
            d.get("usuario", "—") or "—",
            d.get("serial", "—") or "—",
            f"{d.get('fabricante','')} {d.get('modelo','')}".strip() or "—",
            d["score"],
            d["specs_fmt"],
            r["nombre"] if r else "Sin candidato",
            r.get("serial", "—") if r else "—",
            r["score"] if r else "—",
            r["specs_fmt"] if r else "—",
            p["mejora_ram"],
            p["mejora_disco"],
            "Sí" if p["mejora_ssd"] else "No",
            p["ganancia_score"],
        ]
        ws1.append(row)
        ri = ws1.max_row
        for ci in range(1, len(HDR1) + 1):
            c = ws1.cell(ri, ci)
            c.fill = fill(COLOR_DEBIL)
            c.font = cell_font(color=COLOR_WHITE)
            c.border = border()
            c.alignment = left()
        ws1.cell(ri, 5).alignment = center()
        ws1.cell(ri, 9).alignment = center()
        ws1.cell(ri, 11).alignment = center()
        ws1.cell(ri, 12).alignment = center()
        ws1.cell(ri, 13).alignment = center()
        ws1.cell(ri, 14).alignment = center()

    col_widths_1 = [26, 22, 16, 28, 10, 40, 26, 16, 12, 40, 12, 14, 10, 12]
    for ci, w in enumerate(col_widths_1, 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w
    ws1.row_dimensions[1].height = 36

    ws2 = wb.create_sheet("Inventario completo")
    ws2.sheet_view.showGridLines = False

    HDR2 = [
        "Nombre", "Estado", "Usuario", "Serial", "Fabricante", "Modelo",
        "RAM (GB)", "Disco (GB)", "Tipo disco", "CPU", "Score",
    ]
    ws2.append(HDR2)
    for ci, h in enumerate(HDR2, 1):
        cell = ws2.cell(1, ci)
        cell.font = hdr_font(COLOR_ACCENT)
        cell.fill = fill(COLOR_HDR)
        cell.border = border()
        cell.alignment = center()

    for e in sorted(data.get("todos", []), key=lambda x: x.get("score", 0)):
        row = [
            e["nombre"],
            e["estado"],
            e.get("usuario", "—") or "—",
            e.get("serial", "—") or "—",
            e.get("fabricante", "—") or "—",
            e.get("modelo", "—") or "—",
            round(e["ram_mb"] / 1024, 1) if e["ram_mb"] else 0,
            e["disco_gb"],
            e.get("tipo_disco", "—") or "—",
            (e.get("cpu", "—") or "—").split("@")[0].strip(),
            e["score"],
        ]
        ws2.append(row)
        ri = ws2.max_row
        es_activo = "activo" in e["estado"].lower()
        es_debil = e["score"] <= 30
        bg = COLOR_DEBIL if (es_activo and es_debil) else ("0A1F0F" if not es_activo else "111827")
        for ci in range(1, len(HDR2) + 1):
            c = ws2.cell(ri, ci)
            c.fill = fill(bg)
            c.font = cell_font(color=COLOR_WHITE)
            c.border = border()
            c.alignment = left()
        ws2.cell(ri, 7).alignment = center()
        ws2.cell(ri, 8).alignment = center()
        ws2.cell(ri, 11).alignment = center()

    col_widths_2 = [26, 14, 22, 16, 18, 24, 10, 10, 14, 36, 10]
    for ci, w in enumerate(col_widths_2, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.row_dimensions[1].height = 36

    ws3 = wb.create_sheet("Resumen ejecutivo")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 38
    ws3.column_dimensions["B"].width = 18

    resumen = [
        ("Fecha de análisis", date.today().isoformat()),
        ("Total equipos activos", data.get("total_activos", 0)),
        ("Total equipos inactivos", data.get("total_inactivos", 0)),
        ("Activos con specs débiles", data.get("debiles", 0)),
        ("Inactivos aptos para reemplazo", data.get("candidatos", 0)),
        ("Reemplazos recomendados", len([p for p in pares if p["reemplazo"]])),
        ("Sin reemplazo disponible", len([p for p in pares if not p["reemplazo"]])),
    ]
    for label, valor in resumen:
        ws3.append([label, valor])
        ri = ws3.max_row
        ws3.cell(ri, 1).font = cell_font(bold=True, color=COLOR_ACCENT)
        ws3.cell(ri, 1).fill = fill(COLOR_HDR)
        ws3.cell(ri, 1).border = border()
        ws3.cell(ri, 1).alignment = left()
        ws3.cell(ri, 2).font = cell_font(bold=True, color=COLOR_WHITE)
        ws3.cell(ri, 2).fill = fill("1A2235")
        ws3.cell(ri, 2).border = border()
        ws3.cell(ri, 2).alignment = center()
        ws3.row_dimensions[ri].height = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"{filename_prefix}_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/renovacion/excel")
def get_renovacion_excel(modo_prueba: bool = True):
    data = get_renovacion(modo_prueba=modo_prueba)
    return _build_renovacion_excel_response(data, "renovacion_equipos")


@router.post("/renovacion/excel-custom")
def post_renovacion_excel_custom(payload: RenovacionExcelCustomIn):
    data = payload.model_dump()
    return _build_renovacion_excel_response(data, "renovacion_personalizada")


def _aplicar_par_renovacion(
    g: GLPIClient,
    par,
    *,
    estado_reemplazo_id,
    estado_debil_id,
    fecha_iso: str,
    responsable: str,
):
    """Equipo reemplazo: usuario del débil + estado activo. Equipo débil: sin usuario + inactivo."""
    a = par.activo
    r = par.reemplazo
    if not r:
        raise ValueError("Falta equipo de reemplazo")

    uid_debil = a.users_id
    if uid_debil is None:
        uid_debil = 0

    if uid_debil > 0:
        g.update_computer_fields(
            r.id,
            users_id=int(uid_debil),
            states_id=estado_reemplazo_id,
        )
    else:
        g.update_computer_fields(r.id, states_id=estado_reemplazo_id)

    g.update_computer_fields(a.id, users_id=0, states_id=estado_debil_id)

    rp = responsable.strip()
    line = f"{fecha_iso} RENOVACIÓN | {a.nombre} → {r.nombre}"
    if rp:
        line += f" | Resp: {rp}"
    g.append_computer_comment(a.id, line)
    g.append_computer_comment(r.id, line)


@router.post("/renovacion/confirmar")
def post_renovacion_confirmar(payload: RenovacionConfirmarIn):
    to_apply = [p for p in payload.pares if p.reemplazo is not None]
    if not to_apply:
        raise HTTPException(400, "No hay pares con equipo de reemplazo para aplicar.")

    for i, p in enumerate(to_apply):
        rep = p.reemplazo
        if rep is None:
            raise HTTPException(400, f"Par {i + 1}: falta equipo de reemplazo.")
        if not str(p.activo.id).strip():
            raise HTTPException(400, f"Par {i + 1}: falta id del equipo activo.")
        if not str(rep.id).strip():
            raise HTTPException(400, f"Par {i + 1}: falta id del equipo de reemplazo.")

    if payload.modo_prueba:
        aplicados_sim = []
        for p in to_apply:
            rep = p.reemplazo
            assert rep is not None
            aplicados_sim.append({
                "activo_id": str(p.activo.id),
                "activo_nombre": p.activo.nombre,
                "reemplazo_id": str(rep.id),
                "reemplazo_nombre": rep.nombre,
                "detalle": "Simulación: no se modificó GLPI.",
            })
        return {
            "ok": True,
            "modo_prueba": True,
            "aplicados": aplicados_sim,
            "errores": [],
        }

    cfg = get_merged_config()
    if not cfg.get("glpi_url"):
        raise HTTPException(400, "Configura la URL de GLPI primero.")
    if _requests is None:
        raise HTTPException(500, "Instala requests: pip install requests")

    g = GLPIClient(cfg)
    aplicados = []
    errores: list[str] = []
    fecha_iso = date.today().isoformat()

    try:
        g.login()
        activo_sid = g.find_state_id_by_name(payload.estado_reemplazo.strip())
        debil_sid = g.find_state_id_by_name(payload.estado_debil.strip())
        if not activo_sid:
            raise HTTPException(
                400,
                f"No existe un estado «{payload.estado_reemplazo}» en GLPI. "
                "Crea el estado o indica el nombre exacto.",
            )
        if not debil_sid:
            raise HTTPException(
                400,
                f"No existe un estado «{payload.estado_debil}» en GLPI. "
                "Crea el estado o indica el nombre exacto.",
            )

        for p in to_apply:
            rep = p.reemplazo
            assert rep is not None
            try:
                _aplicar_par_renovacion(
                    g,
                    p,
                    estado_reemplazo_id=int(activo_sid),
                    estado_debil_id=int(debil_sid),
                    fecha_iso=fecha_iso,
                    responsable=payload.responsable,
                )
                aplicados.append({
                    "activo_id": str(p.activo.id),
                    "activo_nombre": p.activo.nombre,
                    "reemplazo_id": str(rep.id),
                    "reemplazo_nombre": rep.nombre,
                    "detalle": "Actualizado en GLPI.",
                })
            except HTTPException:
                raise
            except Exception as ex:
                nombre = p.activo.nombre or p.activo.id
                errores.append(f"{nombre}: {ex}")
    except HTTPException:
        raise
    except Exception as ex:
        glpi_http_error(ex, "confirmar renovación en GLPI")
    finally:
        g.logout()

    record_last_glpi_sync()
    return {
        "ok": len(errores) == 0,
        "modo_prueba": False,
        "aplicados": aplicados,
        "errores": errores,
    }
